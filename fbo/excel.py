from openpyxl import load_workbook
import time
import pymysql
from config import host, user_name, password, db_name



def update(text, user):
    try:
        connection = pymysql.connect(
            host=host,
            port=3306,
            user=user_name,
            password=password,
            database=db_name,
            cursorclass=pymysql.cursors.DictCursor
        )

        try:
            with connection.cursor() as cursor:
                cursor.execute(text)
                connection.commit()
                return 0

        finally:
            connection.close()
    except Exception as ex:
        return f'Ошибка {ex}. Не продолжайте!\n\nПередайте руководителю и нажмите <b>ОТМЕНА</b>'

def create(text, user):
    try:
        connection = pymysql.connect(
            host=host,
            port=3306,
            user=user_name,
            password=password,
            database=db_name,
            cursorclass=pymysql.cursors.DictCursor
        )

        try:
            with connection.cursor() as cursor:
                cursor.execute(text)
                connection.commit()
                return 0

        finally:
            connection.close()
    except Exception as ex:
        return f'Ошибка {ex}. Не продолжайте!\n\nПередайте руководителю и нажмите <b>ОТМЕНА</b>'

def selone(text, user):
    try:
        connection = pymysql.connect(
            host=host,
            port=3306,
            user=user_name,
            password=password,
            database=db_name,
            cursorclass=pymysql.cursors.DictCursor
        )


        try:
            with connection.cursor() as cursor:
                cursor.execute(text)
                return cursor.fetchone()

        finally:
            connection.close()
    except Exception as ex:
        return f'Ошибка {ex}. Не продолжайте!\n\nПередайте руководителю и нажмите <b>ОТМЕНА</b>'

def selist(text, user):
    try:
        connection = pymysql.connect(
            host=host,
            port=3306,
            user=user_name,
            password=password,
            database=db_name,
            cursorclass=pymysql.cursors.DictCursor
        )


        try:
            with connection.cursor() as cursor:
                cursor.execute(text)
                return cursor.fetchall()

        finally:
            connection.close()
    except Exception as ex:
        return f'Ошибка {ex}. Не продолжайте!\n\nПередайте руководителю и нажмите <b>ОТМЕНА</b>'

def craate_xl(user, xl_ul_text, xl_tel_text, xl_type_text, xl_count_type_text, xl_mark_text, xl_pack_text, xl_comment_text,
            xl_city_text, xl_count_box_text, xl_count_items_text, xl_comment_city_text, xl_markbox_text):
    fn = 'fbo/zakaz/main_list.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['B1'] = time.strftime("%d.%m.%Y", time.localtime())
    ws['E1'] = user
    ws['B4'] = xl_ul_text
    ws['B5'] = xl_tel_text
    ws['A9'] = xl_type_text
    ws['B9'] = xl_count_type_text
    ws['C9'] = xl_mark_text
    ws['D9'] = xl_pack_text
    ws['E9'] = xl_comment_text
    ws['A31'] = xl_city_text
    ws['B31'] = xl_count_box_text
    ws['C31'] = xl_count_items_text
    ws['D31'] = xl_comment_city_text
    ws['F31'] = xl_markbox_text
    wb.save('fbo/zakaz/zakaz.xlsx')
    wb.close()

def create_kpi(fbo_id, user):
    list_fbo = selone(f"SELECT * FROM fbo WHERE fbo_id = '{fbo_id}'", user)
    this_date = time.strftime("%d.%m.%Y", time.localtime())

    fn = 'checks/' + str(fbo_id) + '.xlsx'
    wb = load_workbook(fn, data_only=True)
    ws = wb['Расчёт стоимости']

    new_fbo = 1
    new_fbo_name = list_fbo['creater']
    # Проверка на брак(визуально) fbo_3_user
    brak_vis = ws['L25'].value
    brak_vis_name = list_fbo['fbo_3_user']
    # Проверка на брак(полный) fbo_3_user
    brak_full = ws['L26'].value
    brak_full_name = list_fbo['fbo_3_user']

    # Бирка для одежды / аксессуара(стандартная) fbo_4_user
    birk = ws['L27'].value
    birk_name = list_fbo['fbo_4_user']

    # Бирка для одежды / аксессуара(брендированная) fbo_4_user
    birk_orig = ws['L28'].value
    birk_orig_name = list_fbo['fbo_4_user']

    # Маркировка товара этикеткой(штрихкод) fbo_4_user fbo_6_user
    mark = ws['L29'].value / 2
    mark_name = list_fbo['fbo_4_user']
    mark_2 = ws['L29'].value / 2
    mark_2_name = list_fbo['fbo_6_user']

    # Макет этикетки ШК fbo_4_user fbo_6_user
    make_mark = ws['L30'].value / 2
    make_mark_name = list_fbo['fbo_4_user']
    make_mark_2 = ws['L30'].value / 2
    make_mark_2_name = list_fbo['fbo_6_user']

    # Сортировка товара(за единицу) fbo_2_user
    sort_items = ws['L32'].value + ws['L31'].value
    sort_items_name = list_fbo['fbo_2_user']

    # Термоусадка fbo_5_user
    term15 = ws['D37'].value
    term25 = ws['D38'].value
    term30 = ws['D39'].value
    term40 = ws['D40'].value
    term60 = ws['D41'].value
    if str(ws['D42'].value).isdigit():
        term_big = ws['D42'].value
    term15_name = list_fbo['fbo_5_user']
    term25_name = list_fbo['fbo_5_user']
    term30_name = list_fbo['fbo_5_user']
    term40_name = list_fbo['fbo_5_user']
    term60_name = list_fbo['fbo_5_user']
    term_big_name = list_fbo['fbo_5_user']

    # Бопп fbo_5_user
    bopp15 = ws['G37'].value
    bopp25 = ws['G38'].value
    bopp30 = ws['G39'].value
    bopp40 = ws['G40'].value
    bopp60 = ws['G41'].value
    if str(ws['G42'].value).isdigit():
        bopp_big = ws['G42'].value
    bopp15_name = list_fbo['fbo_5_user']
    bopp25_name = list_fbo['fbo_5_user']
    bopp30_name = list_fbo['fbo_5_user']
    bopp40_name = list_fbo['fbo_5_user']
    bopp60_name = list_fbo['fbo_5_user']
    bopp_big_name = list_fbo['fbo_5_user']

    # Коробка fbo_5_user
    box15 = ws['J37'].value
    box25 = ws['J38'].value
    box30 = ws['J39'].value
    box40 = ws['J40'].value
    box60 = ws['J41'].value
    if str(ws['J42'].value).isdigit():
        box_big = ws['J42'].value
    box15_name = list_fbo['fbo_5_user']
    box25_name = list_fbo['fbo_5_user']
    box30_name = list_fbo['fbo_5_user']
    box40_name = list_fbo['fbo_5_user']
    box60_name = list_fbo['fbo_5_user']
    box_big_name = list_fbo['fbo_5_user']

    # Стретч - пленка / скотч fbo_5_user
    stretch15 = ws['L37'].value
    stretch25 = ws['L38'].value
    stretch30 = ws['L39'].value
    stretch40 = ws['L40'].value
    stretch60 = ws['L41'].value
    if str(ws['L42'].value).isdigit():
        stretch_big = ws['L42'].value
    stretch15_name = list_fbo['fbo_5_user']
    stretch25_name = list_fbo['fbo_5_user']
    stretch30_name = list_fbo['fbo_5_user']
    stretch40_name = list_fbo['fbo_5_user']
    stretch60_name = list_fbo['fbo_5_user']
    stretch_big_name = list_fbo['fbo_5_user']

    # Пупырка fbo_5_user
    bubble15 = ws['D47'].value
    bubble25 = ws['D48'].value
    bubble30 = ws['D49'].value
    bubble40 = ws['D50'].value
    bubble60 = ws['D51'].value
    if str(ws['D52'].value).isdigit():
        bubble_big = ws['D52'].value
    bubble15_name = list_fbo['fbo_5_user']
    bubble25_name = list_fbo['fbo_5_user']
    bubble30_name = list_fbo['fbo_5_user']
    bubble40_name = list_fbo['fbo_5_user']
    bubble60_name = list_fbo['fbo_5_user']
    bubble_big_name = list_fbo['fbo_5_user']

    # Зип пакет fbo_5_user
    zip15 = ws['G47'].value
    zip25 = ws['G48'].value
    zip30 = ws['G49'].value
    zip40 = ws['G50'].value
    zip60 = ws['G51'].value
    if str(ws['G52'].value).isdigit():
        zip_big = ws['G52'].value
    zip15_name = list_fbo['fbo_5_user']
    zip25_name = list_fbo['fbo_5_user']
    zip30_name = list_fbo['fbo_5_user']
    zip40_name = list_fbo['fbo_5_user']
    zip60_name = list_fbo['fbo_5_user']
    zip_big_name = list_fbo['fbo_5_user']

    # Курьерский пакет fbo_5_user
    curpack15 = ws['J47'].value
    curpack25 = ws['J48'].value
    curpack30 = ws['J49'].value
    curpack40 = ws['J50'].value
    curpack60 = ws['J51'].value
    if str(ws['J52'].value).isdigit():
        curpack_big = ws['J52'].value
    curpack15_name = list_fbo['fbo_5_user']
    curpack25_name = list_fbo['fbo_5_user']
    curpack30_name = list_fbo['fbo_5_user']
    curpack40_name = list_fbo['fbo_5_user']
    curpack60_name = list_fbo['fbo_5_user']
    curpack_big_name = list_fbo['fbo_5_user']

    # Под запайку fbo_5_user
    zap15 = ws['L47'].value
    zap25 = ws['L48'].value
    zap30 = ws['L49'].value
    zap40 = ws['L50'].value
    zap60 = ws['L51'].value
    if str(ws['L42'].value).isdigit():
        zap_big = ws['L52'].value
    zap15_name = list_fbo['fbo_5_user']
    zap25_name = list_fbo['fbo_5_user']
    zap30_name = list_fbo['fbo_5_user']
    zap40_name = list_fbo['fbo_5_user']
    zap60_name = list_fbo['fbo_5_user']
    zap_big_name = list_fbo['fbo_5_user']

    # Зип-пакет с бегунком fbo_5_user
    ziprun15 = ws['D57'].value
    ziprun25 = ws['D58'].value
    ziprun30 = ws['D59'].value
    ziprun40 = ws['D60'].value
    ziprun60 = ws['D61'].value
    if str(ws['D42'].value).isdigit():
        ziprun_big = ws['D62'].value
    ziprun15_name = list_fbo['fbo_5_user']
    ziprun25_name = list_fbo['fbo_5_user']
    ziprun30_name = list_fbo['fbo_5_user']
    ziprun40_name = list_fbo['fbo_5_user']
    ziprun60_name = list_fbo['fbo_5_user']
    ziprun_big_name = list_fbo['fbo_5_user']

    # Сбор товаров в гофрокороб fbo_7_user
    sort_box = ws['L65'].value
    sort_box_name = list_fbo['fbo_7_user']

    # Сбор товаров на паллету fbo_7_user
    sort_pal = ws['L66'].value
    sort_pal_name = list_fbo['fbo_7_user']

    # Маркировка коробов (ШК ТТН, ШК короба) fbo_9_user
    mark_box = ws['L67'].value
    mark_box_name = list_fbo['fbo_9_user']

    # Паллетирование палеты стретч-пленкой fbo_10_user
    pal = ws['L68'].value
    pal_name = list_fbo['fbo_10_user']

    # Формирование накладной на отправку до 10 артикулов fbo_8_user
    post = ws['K69'].value
    dop_post = ws['L69'].value
    post_name = list_fbo['fbo_8_user']
    dop_post_name = list_fbo['fbo_8_user']

    # Доставка короба fbo_17_user
    dost_box = ws['E75'].value + ws['E76'].value + ws['E77'].value + ws['E78'].value + ws['E80'].value + ws['E81'].value + ws['E82'].value + ws['P75'].value + ws['P76'].value + ws['P77'].value + ws['P78'].value + ws['P79'].value + ws['P80'].value + ws['P81'].value + ws['P82'].value
    if list_fbo['fbo_17_user'] == '':
        dost_box_name = list_fbo['fbo_19_user']
    else:
        dost_box_name = list_fbo['fbo_17_user']

    # Доставка паллеты fbo_17_user
    dost_pal = ws['K75'].value + ws['K76'].value + ws['K77'].value + ws['K78'].value + ws['K79'].value + ws['K80'].value + ws['K81'].value + ws['K82'].value
    if list_fbo['fbo_17_user'] == '':
        dost_pal_name = list_fbo['fbo_19_user']
    else:
        dost_pal_name = list_fbo['fbo_17_user']

    # Гофрокороб 60х40х40 fbo_7_user
    box_60_40 = ws['L85'].value
    box_60_40_name = list_fbo['fbo_7_user']

    # Гофрокороб до 60 см fbo_7_user
    box_small = ws['L86'].value
    box_small_name = list_fbo['fbo_7_user']

    # Гофрокороб более 60 см fbo_7_user
    box_big_2 = ws['L87'].value
    box_big_2_name = list_fbo['fbo_7_user']

    create(f"REPLACE INTO kpi(id_fbo, date_fbo, done, new_fbo, new_fbo_name, brak_vis, brak_vis_name, brak_full, "
           f"brak_full_name, birk, birk_name, birk_orig, birk_orig_name, mark, mark_name, mark_2, mark_2_name, "
           f"make_mark, make_mark_name, make_mark_2, make_mark_2_name, sort_items, sort_items_name, term15, term15_name, "
           f"term25, term25_name, term30, term30_name, term40, term40_name, term60, term60_name, term_big, term_big_name, "
           f"bopp15, bopp15_name, bopp25, bopp25_name, bopp30, bopp30_name, bopp40, bopp40_name, bopp60, bopp60_name, "
           f"bopp_big, bopp_big_name, box15, box15_name, box25, box25_name, box30, box30_name, box40, box40_name, box60, "
           f"box60_name, box_big, box_big_name, stretch15, stretch15_name, stretch25, stretch25_name, stretch30, "
           f"stretch30_name, stretch40, stretch40_name, stretch60, stretch60_name, stretch_big, stretch_big_name, "
           f"bubble15, bubble15_name, bubble25, bubble25_name, bubble30, bubble30_name, bubble40, bubble40_name, "
           f"bubble60, bubble60_name, bubble_big, bubble_big_name, zip15, zip15_name, zip25, zip25_name, zip30, "
           f"zip30_name, zip40, zip40_name, zip60, zip60_name, zip_big, zip_big_name, curpack15, curpack15_name, "
           f"curpack25, curpack25_name, curpack30, curpack30_name, curpack40, curpack40_name, curpack60, "
           f"curpack60_name, curpack_big, curpack_big_name, zap15, zap15_name, zap25, zap25_name, zap30, zap30_name, "
           f"zap40, zap40_name, zap60, zap60_name, zap_big, zap_big_name, ziprun15, ziprun15_name, ziprun25, "
           f"ziprun25_name, ziprun30, ziprun30_name, ziprun40, ziprun40_name, ziprun60, ziprun60_name, ziprun_big, "
           f"ziprun_big_name, sort_box, sort_box_name, sort_pal, sort_pal_name, mark_box, mark_box_name, pal, pal_name, "
           f"post, post_name, dop_post, dop_post_name, dost_box, dost_box_name, dost_pal, dost_pal_name, box_60_40, "
           f"box_60_40_name, box_small, box_small_name, box_big_2, box_big_2_name, ff_city) "
           f"VALUES ('{fbo_id}', '{this_date}', '{list_fbo['done']}', '{new_fbo}', '{new_fbo_name}', '{brak_vis}', "
           f"'{brak_vis_name}', '{brak_full}', '{brak_full_name}', '{birk}', '{birk_name}', '{birk_orig}', "
           f"'{birk_orig_name}', '{mark}', '{mark_name}', '{mark_2}', '{mark_2_name}', '{make_mark}', "
           f"'{make_mark_name}', '{make_mark_2}', '{make_mark_2_name}', '{sort_items}', '{sort_items_name}', "
           f"'{term15}', '{term15_name}', '{term25}', '{term25_name}', '{term30}', '{term30_name}', '{term40}', "
           f"'{term40_name}', '{term60}', '{term60_name}', '{term_big}', '{term_big_name}', '{bopp15}', "
           f"'{bopp15_name}', '{bopp25}', '{bopp25_name}', '{bopp30}', '{bopp30_name}', '{bopp40}', '{bopp40_name}', "
           f"'{bopp60}', '{bopp60_name}', '{bopp_big}', '{bopp_big_name}', '{box15}', '{box15_name}', '{box25}', "
           f"'{box25_name}', '{box30}', '{box30_name}', '{box40}', '{box40_name}', '{box60}', '{box60_name}', "
           f"'{box_big}', '{box_big_name}', '{stretch15}', '{stretch15_name}', '{stretch25}', '{stretch25_name}', "
           f"'{stretch30}', '{stretch30_name}', '{stretch40}', '{stretch40_name}', '{stretch60}', '{stretch60_name}', "
           f"'{stretch_big}', '{stretch_big_name}', '{bubble15}', '{bubble15_name}', '{bubble25}', '{bubble25_name}', "
           f"'{bubble30}', '{bubble30_name}', '{bubble40}', '{bubble40_name}', '{bubble60}', '{bubble60_name}', "
           f"'{bubble_big}', '{bubble_big_name}', '{zip15}', '{zip15_name}', '{zip25}', '{zip25_name}', '{zip30}', "
           f"'{zip30_name}', '{zip40}', '{zip40_name}', '{zip60}', '{zip60_name}', '{zip_big}', '{zip_big_name}', "
           f"'{curpack15}', '{curpack15_name}', '{curpack25}', '{curpack25_name}', '{curpack30}', '{curpack30_name}', "
           f"'{curpack40}', '{curpack40_name}', '{curpack60}', '{curpack60_name}', '{curpack_big}', '{curpack_big_name}', "
           f"'{zap15}', '{zap15_name}', '{zap25}', '{zap25_name}', '{zap30}', '{zap30_name}', '{zap40}', '{zap40_name}', "
           f"'{zap60}', '{zap60_name}', '{zap_big}', '{zap_big_name}', '{ziprun15}', '{ziprun15_name}', '{ziprun25}', "
           f"'{ziprun25_name}', '{ziprun30}', '{ziprun30_name}', '{ziprun40}', '{ziprun40_name}', '{ziprun60}', "
           f"'{ziprun60_name}', '{ziprun_big}', '{ziprun_big_name}', '{sort_box}', '{sort_box_name}', '{sort_pal}', "
           f"'{sort_pal_name}', '{mark_box}', '{mark_box_name}', '{pal}', '{pal_name}', '{post}', '{post_name}', "
           f"'{dop_post}', '{dop_post_name}', '{dost_box}', '{dost_box_name}', '{dost_pal}', '{dost_pal_name}', "
           f"'{box_60_40}', '{box_60_40_name}', '{box_small}', '{box_small_name}', '{box_big_2}', "
           f"'{box_big_2_name}', 0)", user)
    wb.close()

def create_kpi_pd(fbo_id, message, user):
    list_fbo = selone(f"SELECT * FROM fbo WHERE fbo_id = '{fbo_id}'", user)
    this_date = time.strftime("%d.%m.%Y", time.localtime())

    new_fbo = 1
    new_fbo_name = list_fbo['creater']

    # Проверка на брак(визуально) fbo_3_user
    brak_vis = 0
    brak_vis_name = list_fbo['fbo_3_user']

    # Проверка на брак(полный) fbo_3_user
    brak_full = 0
    brak_full_name = list_fbo['fbo_3_user']

    # Бирка для одежды / аксессуара(стандартная) fbo_4_user
    birk = 0
    birk_name = list_fbo['fbo_4_user']

    # Бирка для одежды / аксессуара(брендированная) fbo_4_user
    birk_orig = 0
    birk_orig_name = list_fbo['fbo_4_user']

    # Маркировка товара этикеткой(штрихкод) fbo_4_user fbo_6_user
    mark = 0
    mark_name = list_fbo['fbo_4_user']
    mark_2 = 0
    mark_2_name = list_fbo['fbo_6_user']

    # Макет этикетки ШК fbo_4_user fbo_6_user
    make_mark = 0
    make_mark_name = list_fbo['fbo_4_user']
    make_mark_2 = 0
    make_mark_2_name = list_fbo['fbo_6_user']

    # Сортировка товара(за единицу) fbo_2_user
    sort_items = 0
    sort_items_name = list_fbo['fbo_2_user']

    # Термоусадка fbo_5_user
    term15 = 0
    term25 = 0
    term30 = 0
    term40 = 0
    term60 = 0
    term_big = 0
    term15_name = list_fbo['fbo_5_user']
    term25_name = list_fbo['fbo_5_user']
    term30_name = list_fbo['fbo_5_user']
    term40_name = list_fbo['fbo_5_user']
    term60_name = list_fbo['fbo_5_user']
    term_big_name = list_fbo['fbo_5_user']

    # Бопп fbo_5_user
    bopp15 = 0
    bopp25 = 0
    bopp30 = 0
    bopp40 = 0
    bopp60 = 0
    bopp_big = 0
    bopp15_name = list_fbo['fbo_5_user']
    bopp25_name = list_fbo['fbo_5_user']
    bopp30_name = list_fbo['fbo_5_user']
    bopp40_name = list_fbo['fbo_5_user']
    bopp60_name = list_fbo['fbo_5_user']
    bopp_big_name = list_fbo['fbo_5_user']

    # Коробка fbo_5_user
    box15 = 0
    box25 = 0
    box30 = 0
    box40 = 0
    box60 = 0
    box_big = 0
    box15_name = list_fbo['fbo_5_user']
    box25_name = list_fbo['fbo_5_user']
    box30_name = list_fbo['fbo_5_user']
    box40_name = list_fbo['fbo_5_user']
    box60_name = list_fbo['fbo_5_user']
    box_big_name = list_fbo['fbo_5_user']

    # Стретч - пленка / скотч fbo_5_user
    stretch15 = 0
    stretch25 = 0
    stretch30 = 0
    stretch40 = 0
    stretch60 = 0
    stretch_big = 0
    stretch15_name = list_fbo['fbo_5_user']
    stretch25_name = list_fbo['fbo_5_user']
    stretch30_name = list_fbo['fbo_5_user']
    stretch40_name = list_fbo['fbo_5_user']
    stretch60_name = list_fbo['fbo_5_user']
    stretch_big_name = list_fbo['fbo_5_user']

    # Пупырка fbo_5_user
    bubble15 = 0
    bubble25 = 0
    bubble30 = 0
    bubble40 = 0
    bubble60 = 0
    bubble_big = 0
    bubble15_name = list_fbo['fbo_5_user']
    bubble25_name = list_fbo['fbo_5_user']
    bubble30_name = list_fbo['fbo_5_user']
    bubble40_name = list_fbo['fbo_5_user']
    bubble60_name = list_fbo['fbo_5_user']
    bubble_big_name = list_fbo['fbo_5_user']

    # Зип пакет fbo_5_user
    zip15 = 0
    zip25 = 0
    zip30 = 0
    zip40 = 0
    zip60 = 0
    zip_big = 0
    zip15_name = list_fbo['fbo_5_user']
    zip25_name = list_fbo['fbo_5_user']
    zip30_name = list_fbo['fbo_5_user']
    zip40_name = list_fbo['fbo_5_user']
    zip60_name = list_fbo['fbo_5_user']
    zip_big_name = list_fbo['fbo_5_user']

    # Курьерский пакет fbo_5_user
    curpack15 = 0
    curpack25 = 0
    curpack30 = 0
    curpack40 = 0
    curpack60 = 0
    curpack_big = 0
    curpack15_name = list_fbo['fbo_5_user']
    curpack25_name = list_fbo['fbo_5_user']
    curpack30_name = list_fbo['fbo_5_user']
    curpack40_name = list_fbo['fbo_5_user']
    curpack60_name = list_fbo['fbo_5_user']
    curpack_big_name = list_fbo['fbo_5_user']

    # Под запайку fbo_5_user
    zap15 = 0
    zap25 = 0
    zap30 = 0
    zap40 = 0
    zap60 = 0
    zap_big = 0
    zap15_name = list_fbo['fbo_5_user']
    zap25_name = list_fbo['fbo_5_user']
    zap30_name = list_fbo['fbo_5_user']
    zap40_name = list_fbo['fbo_5_user']
    zap60_name = list_fbo['fbo_5_user']
    zap_big_name = list_fbo['fbo_5_user']

    # Зип-пакет с бегунком fbo_5_user
    ziprun15 = 0
    ziprun25 = 0
    ziprun30 = 0
    ziprun40 = 0
    ziprun60 = 0
    ziprun_big = 0
    ziprun15_name = list_fbo['fbo_5_user']
    ziprun25_name = list_fbo['fbo_5_user']
    ziprun30_name = list_fbo['fbo_5_user']
    ziprun40_name = list_fbo['fbo_5_user']
    ziprun60_name = list_fbo['fbo_5_user']
    ziprun_big_name = list_fbo['fbo_5_user']

    # Сбор товаров в гофрокороб fbo_7_user
    sort_box = 0
    sort_box_name = list_fbo['fbo_7_user']

    # Сбор товаров на паллету fbo_7_user
    sort_pal = 0
    sort_pal_name = list_fbo['fbo_7_user']

    # Маркировка коробов (ШК ТТН, ШК короба) fbo_9_user
    mark_box = 0
    mark_box_name = list_fbo['fbo_9_user']

    # Паллетирование палеты стретч-пленкой fbo_10_user
    pal = 0
    pal_name = list_fbo['fbo_10_user']

    # Формирование накладной на отправку до 10 артикулов fbo_8_user
    post = 0
    dop_post = 0
    post_name = list_fbo['fbo_8_user']
    dop_post_name = list_fbo['fbo_8_user']

    # Доставка короба fbo_17_user
    dost_box = message
    if list_fbo['fbo_17_user'] == '':
        dost_box_name = list_fbo['fbo_19_user']
    else:
        dost_box_name = list_fbo['fbo_17_user']

    # Доставка паллеты fbo_17_user
    dost_pal = 0
    if list_fbo['fbo_17_user'] == '':
        dost_pal_name = list_fbo['fbo_19_user']
    else:
        dost_pal_name = list_fbo['fbo_17_user']

    # Гофрокороб 60х40х40 fbo_7_user
    box_60_40 = 0
    box_60_40_name = list_fbo['fbo_7_user']

    # Гофрокороб до 60 см fbo_7_user
    box_small = 0
    box_small_name = list_fbo['fbo_7_user']

    # Гофрокороб более 60 см fbo_7_user
    box_big_2 = 0
    box_big_2_name = list_fbo['fbo_7_user']

    create(f"REPLACE INTO kpi(id_fbo, date_fbo, done, new_fbo, new_fbo_name, brak_vis, brak_vis_name, brak_full, "
           f"brak_full_name, birk, birk_name, birk_orig, birk_orig_name, mark, mark_name, mark_2, mark_2_name, "
           f"make_mark, make_mark_name, make_mark_2, make_mark_2_name, sort_items, sort_items_name, term15, term15_name, "
           f"term25, term25_name, term30, term30_name, term40, term40_name, term60, term60_name, term_big, term_big_name, "
           f"bopp15, bopp15_name, bopp25, bopp25_name, bopp30, bopp30_name, bopp40, bopp40_name, bopp60, bopp60_name, "
           f"bopp_big, bopp_big_name, box15, box15_name, box25, box25_name, box30, box30_name, box40, box40_name, box60, "
           f"box60_name, box_big, box_big_name, stretch15, stretch15_name, stretch25, stretch25_name, stretch30, "
           f"stretch30_name, stretch40, stretch40_name, stretch60, stretch60_name, stretch_big, stretch_big_name, "
           f"bubble15, bubble15_name, bubble25, bubble25_name, bubble30, bubble30_name, bubble40, bubble40_name, "
           f"bubble60, bubble60_name, bubble_big, bubble_big_name, zip15, zip15_name, zip25, zip25_name, zip30, "
           f"zip30_name, zip40, zip40_name, zip60, zip60_name, zip_big, zip_big_name, curpack15, curpack15_name, "
           f"curpack25, curpack25_name, curpack30, curpack30_name, curpack40, curpack40_name, curpack60, "
           f"curpack60_name, curpack_big, curpack_big_name, zap15, zap15_name, zap25, zap25_name, zap30, zap30_name, "
           f"zap40, zap40_name, zap60, zap60_name, zap_big, zap_big_name, ziprun15, ziprun15_name, ziprun25, "
           f"ziprun25_name, ziprun30, ziprun30_name, ziprun40, ziprun40_name, ziprun60, ziprun60_name, ziprun_big, "
           f"ziprun_big_name, sort_box, sort_box_name, sort_pal, sort_pal_name, mark_box, mark_box_name, pal, pal_name, "
           f"post, post_name, dop_post, dop_post_name, dost_box, dost_box_name, dost_pal, dost_pal_name, box_60_40, "
           f"box_60_40_name, box_small, box_small_name, box_big_2, box_big_2_name, ff_city) "
           f"VALUES ('{fbo_id}', '{this_date}', '{list_fbo['done']}', '{new_fbo}', '{new_fbo_name}', '{brak_vis}', "
           f"'{brak_vis_name}', '{brak_full}', '{brak_full_name}', '{birk}', '{birk_name}', '{birk_orig}', "
           f"'{birk_orig_name}', '{mark}', '{mark_name}', '{mark_2}', '{mark_2_name}', '{make_mark}', "
           f"'{make_mark_name}', '{make_mark_2}', '{make_mark_2_name}', '{sort_items}', '{sort_items_name}', "
           f"'{term15}', '{term15_name}', '{term25}', '{term25_name}', '{term30}', '{term30_name}', '{term40}', "
           f"'{term40_name}', '{term60}', '{term60_name}', '{term_big}', '{term_big_name}', '{bopp15}', "
           f"'{bopp15_name}', '{bopp25}', '{bopp25_name}', '{bopp30}', '{bopp30_name}', '{bopp40}', '{bopp40_name}', "
           f"'{bopp60}', '{bopp60_name}', '{bopp_big}', '{bopp_big_name}', '{box15}', '{box15_name}', '{box25}', "
           f"'{box25_name}', '{box30}', '{box30_name}', '{box40}', '{box40_name}', '{box60}', '{box60_name}', "
           f"'{box_big}', '{box_big_name}', '{stretch15}', '{stretch15_name}', '{stretch25}', '{stretch25_name}', "
           f"'{stretch30}', '{stretch30_name}', '{stretch40}', '{stretch40_name}', '{stretch60}', '{stretch60_name}', "
           f"'{stretch_big}', '{stretch_big_name}', '{bubble15}', '{bubble15_name}', '{bubble25}', '{bubble25_name}', "
           f"'{bubble30}', '{bubble30_name}', '{bubble40}', '{bubble40_name}', '{bubble60}', '{bubble60_name}', "
           f"'{bubble_big}', '{bubble_big_name}', '{zip15}', '{zip15_name}', '{zip25}', '{zip25_name}', '{zip30}', "
           f"'{zip30_name}', '{zip40}', '{zip40_name}', '{zip60}', '{zip60_name}', '{zip_big}', '{zip_big_name}', "
           f"'{curpack15}', '{curpack15_name}', '{curpack25}', '{curpack25_name}', '{curpack30}', '{curpack30_name}', "
           f"'{curpack40}', '{curpack40_name}', '{curpack60}', '{curpack60_name}', '{curpack_big}', '{curpack_big_name}', "
           f"'{zap15}', '{zap15_name}', '{zap25}', '{zap25_name}', '{zap30}', '{zap30_name}', '{zap40}', '{zap40_name}', "
           f"'{zap60}', '{zap60_name}', '{zap_big}', '{zap_big_name}', '{ziprun15}', '{ziprun15_name}', '{ziprun25}', "
           f"'{ziprun25_name}', '{ziprun30}', '{ziprun30_name}', '{ziprun40}', '{ziprun40_name}', '{ziprun60}', "
           f"'{ziprun60_name}', '{ziprun_big}', '{ziprun_big_name}', '{sort_box}', '{sort_box_name}', '{sort_pal}', "
           f"'{sort_pal_name}', '{mark_box}', '{mark_box_name}', '{pal}', '{pal_name}', '{post}', '{post_name}', "
           f"'{dop_post}', '{dop_post_name}', '{dost_box}', '{dost_box_name}', '{dost_pal}', '{dost_pal_name}', "
           f"'{box_60_40}', '{box_60_40_name}', '{box_small}', '{box_small_name}', '{box_big_2}', "
           f"'{box_big_2_name}', 0)", user)


def call_ul(user, list_ul):
    fn = 'fbo/call.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    count = 3
    for name_ul, value_ul in list_ul.items():
        ws[f'A{count}'] = value_ul[0]
        ws[f'D{count}'] = value_ul[1]
        ws[f'O{count}'] = name_ul
        ws[f'X{count}'] = value_ul[2]
        count += 1

    wb.save(f'fbo/{user}.xlsx')
    wb.close()
